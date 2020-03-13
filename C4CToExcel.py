from tkinter import filedialog
from tkinter import *
import openpyxl as xl

#root = Tk()
#root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("C4C Excel File","*.xlsx"),("all files","*.*")))
filename = "c://Users//ptellapaneni//Downloads//SAPAnalyticsReport(Z4DF48504FDA71F4AF36D5).xlsx"
wb = xl.load_workbook(filename = filename)
ws = wb.active


# loop till the first cell is geomarket and the second is geomarket / sub geomarket
for i in range(1, ws.max_row):
    if str(ws.cell(row = i, column=1).value).find("GeoMarket") !=-1 and str(ws.cell(row = i, column=2).value).find("GeoMarket") !=-1:
        rowWithData = i+1
        HeaderDataStart = i-1
        break

# Now Read all the relevant Columns into the relevant lists




